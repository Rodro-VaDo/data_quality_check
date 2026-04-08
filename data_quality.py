"""
calidad_mes_anterior_participantes_servicios.py
===============================================
Evalúa la completitud mínima y la coherencia lógica de los registros de un
período en la base all_participants_service_families_data, hojas
'participantes' y 'servicios'.

Campos mínimos evaluados
------------------------
PARTICIPANTES (siempre):
  TIPO PARTICIPANTE, NOMBRES Y APELLIDOS DEL PARTICIPANTE,
  TIPO DE DOCUMENTO, NUMERO DE DOCUMENTO, FECHA DE NACIMIENTO,
  SEXO DEL PARTICIPANTE, GENERO DEL PARTICIPANTE,
  DEPARTAMENTO (RESIDENCIA), MUNICIPIO (RESIDENCIA)

PARTICIPANTES (solo si TIPO PARTICIPANTE contiene 'PARD'):
  DEFENSOR/A, FECHA APERTURA PARD

SERVICIOS:
  PROGRAMA, GRUPO DE SERVICIO (FEDERATIVOS), SERVICIO (FEDERATIVOS),
  SUB-SERVICIO, FAMILIA DE ORIGEN (*), FAMILIA / CASA DE ACOGIDA,
  FECHA DE INGRESO A ALDEAS, FECHA DE ENTRADA AL SERVICIO,
  MOTIVO DE ENTRADA AL SERVICIO
  (*) FAMILIA DE ORIGEN en FFC/FLC: solo advertencia, nunca error de completitud.
  (**) FAMILIA / CASA DE ACOGIDA en DFE: nunca requerida (DFE trabaja con familia de ORIGEN).

Validaciones lógicas
--------------------
  1. Fecha nacimiento: no futura, >= 1920.
  2. Edad declarada vs. calculada desde nacimiento (±1 año).
  3. Grupo etario vs. edad.
  4. Tipo documento vs. formato número; fallback a nombre normalizado
     para detectar duplicados cuando el número está vacío.
  5. Duplicados de participante por (tipo+número) o nombre normalizado.
  6. Fecha apertura PARD: no futura, posterior a nacimiento (solo PARD).
  7. Fechas de servicio no futuras.
  8. FECHA INGRESO ALDEAS <= FECHA ENTRADA SERVICIO.
  9. FECHA SALIDA >= FECHA ENTRADA SERVICIO.
 10. DFE con FAMILIA ACOGIDA diligenciada: alerta crítica modelo relacional.
     FFC/FLC sin FAMILIA ORIGEN: advertencia.
 11. Duplicados de ID SERVICIO.
 12. Servicios huérfanos (ID participante inexistente) y
     participantes huérfanos (sin ningún servicio).
 13. Fecha entrada servicio posterior a fecha de nacimiento.
 14. Tipo participante vs. grupo de servicio (advertencia).

Secciones
---------
1  Constantes de UI
2  Diálogo de selección de período   ← idéntico a estadísticas_2025
3  GUI de Progreso                   ← idéntico a estadísticas_2025
4  Carga y normalización
5  Utilidades de texto y completitud
6  Mapas de campos mínimos y constantes de validación lógica
7  Completitud (máscaras de presencia)
8  Validaciones lógicas
9  Pipeline principal + resúmenes    ← estructura idéntica a estadísticas_2025
10 Escritura del reporte             ← estructura idéntica a estadísticas_2025
11 Punto de entrada (main)           ← idéntico a estadísticas_2025
"""

from __future__ import annotations

import re
import threading
import queue
import unicodedata
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd

import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# ============================================================
# SECCIÓN 1 — CONSTANTES DE UI
# ============================================================

MESES_NOMBRE: List[str] = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


# ============================================================
# SECCIÓN 2 — DIÁLOGO DE SELECCIÓN DE PERÍODO
# ============================================================

class PeriodSelectorDialog:
    """
    Diálogo modal para seleccionar el período a evaluar.
    Idéntico al de comprobador_estadisticas_2025.py.

    result : Tuple[int, int, int, int] | None
        (mes_ini, anio_ini, mes_fin, anio_fin) si el usuario confirmó,
        None si canceló o cerró la ventana.
    """

    def __init__(self, parent: tk.Tk):
        self.result: Optional[Tuple[int, int, int, int]] = None

        self.top = tk.Toplevel(parent)
        self.top.title("Seleccionar período de evaluación")
        self.top.resizable(False, False)
        self.top.grab_set()
        self.top.protocol("WM_DELETE_WINDOW", self._cancelar)

        anio_actual = datetime.now().year
        anios = [str(a) for a in range(2020, anio_actual + 2)]

        self._modo     = tk.StringVar(value="mes")
        self._mes_ini  = tk.StringVar()
        self._anio_ini = tk.StringVar()
        self._mes_fin  = tk.StringVar()
        self._anio_fin = tk.StringVar()

        pad = {"padx": 14, "pady": 6}

        frm_modo = ttk.LabelFrame(self.top, text="Modo de período")
        frm_modo.pack(fill="x", **pad)
        ttk.Radiobutton(frm_modo, text="Mes específico",
                        variable=self._modo, value="mes",
                        command=self._update_ui).pack(side="left", padx=10, pady=6)
        ttk.Radiobutton(frm_modo, text="Rango de meses",
                        variable=self._modo, value="rango",
                        command=self._update_ui).pack(side="left", padx=10, pady=6)

        self._frm_campos = ttk.Frame(self.top)
        self._frm_campos.pack(fill="x", padx=14, pady=4)

        self._frm_ini = ttk.LabelFrame(self._frm_campos, text="Mes / Año")
        self._frm_ini.grid(row=0, column=0, padx=(0, 8), pady=4, sticky="nsew")
        self._cb_mes_ini  = ttk.Combobox(self._frm_ini, textvariable=self._mes_ini,
                                          values=MESES_NOMBRE, state="readonly", width=13)
        self._cb_mes_ini.grid(row=0, column=0, padx=8, pady=8)
        self._cb_anio_ini = ttk.Combobox(self._frm_ini, textvariable=self._anio_ini,
                                          values=anios, state="readonly", width=7)
        self._cb_anio_ini.grid(row=0, column=1, padx=8, pady=8)

        self._frm_fin = ttk.LabelFrame(self._frm_campos, text="Hasta (Mes / Año)")
        self._frm_fin.grid(row=0, column=1, padx=(8, 0), pady=4, sticky="nsew")
        self._cb_mes_fin  = ttk.Combobox(self._frm_fin, textvariable=self._mes_fin,
                                          values=MESES_NOMBRE, state="readonly", width=13)
        self._cb_mes_fin.grid(row=0, column=0, padx=8, pady=8)
        self._cb_anio_fin = ttk.Combobox(self._frm_fin, textvariable=self._anio_fin,
                                          values=anios, state="readonly", width=7)
        self._cb_anio_fin.grid(row=0, column=1, padx=8, pady=8)

        frm_btn = ttk.Frame(self.top)
        frm_btn.pack(fill="x", padx=14, pady=(4, 14))
        ttk.Button(frm_btn, text="Cancelar",
                   command=self._cancelar).pack(side="right", padx=(6, 0))
        ttk.Button(frm_btn, text="Continuar →",
                   command=self._confirmar).pack(side="right")

        self._update_ui()
        self._centrar(parent)
        parent.wait_window(self.top)

    def _centrar(self, parent: tk.Tk) -> None:
        self.top.update_idletasks()
        px, py = parent.winfo_x(), parent.winfo_y()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        tw, th = self.top.winfo_width(), self.top.winfo_height()
        self.top.geometry(f"+{px + (pw - tw) // 2}+{py + (ph - th) // 2}")

    def _update_ui(self) -> None:
        if self._modo.get() == "mes":
            self._frm_ini.config(text="Mes / Año a evaluar")
            self._frm_fin.grid_remove()
        else:
            self._frm_ini.config(text="Desde (Mes / Año)")
            self._frm_fin.grid()

    def _cancelar(self) -> None:
        self.result = None
        self.top.destroy()

    def _confirmar(self) -> None:
        mes_ini_str  = self._mes_ini.get()
        anio_ini_str = self._anio_ini.get()
        if not mes_ini_str or not anio_ini_str:
            messagebox.showwarning("Dato faltante",
                "Selecciona el mes y año.", parent=self.top)
            return
        mes_ini  = MESES_NOMBRE.index(mes_ini_str) + 1
        anio_ini = int(anio_ini_str)
        if self._modo.get() == "mes":
            mes_fin  = mes_ini
            anio_fin = anio_ini
        else:
            mes_fin_str  = self._mes_fin.get()
            anio_fin_str = self._anio_fin.get()
            if not mes_fin_str or not anio_fin_str:
                messagebox.showwarning("Dato faltante",
                    "Selecciona el mes y año de fin del rango.", parent=self.top)
                return
            mes_fin  = MESES_NOMBRE.index(mes_fin_str) + 1
            anio_fin = int(anio_fin_str)
            if (anio_fin, mes_fin) < (anio_ini, mes_ini):
                messagebox.showwarning("Rango inválido",
                    "El mes/año de fin debe ser igual o posterior al de inicio.",
                    parent=self.top)
                return
        self.result = (mes_ini, anio_ini, mes_fin, anio_fin)
        self.top.destroy()


def period_label(mes_ini: int, anio_ini: int, mes_fin: int, anio_fin: int) -> str:
    if (mes_ini, anio_ini) == (mes_fin, anio_fin):
        return f"{anio_ini}_{mes_ini:02d}"
    return f"{anio_ini}_{mes_ini:02d}_a_{anio_fin}_{mes_fin:02d}"


def period_display(mes_ini: int, anio_ini: int, mes_fin: int, anio_fin: int) -> str:
    ini = f"{MESES_NOMBRE[mes_ini - 1]} {anio_ini}"
    fin = f"{MESES_NOMBRE[mes_fin - 1]} {anio_fin}"
    return ini if ini == fin else f"{ini} → {fin}"


def period_bounds(
    mes_ini: int, anio_ini: int, mes_fin: int, anio_fin: int
) -> Tuple[pd.Timestamp, pd.Timestamp]:
    start = pd.Timestamp(date(anio_ini, mes_ini, 1)).normalize()
    if mes_fin == 12:
        last_day = pd.Timestamp(date(anio_fin + 1, 1, 1)) - pd.Timedelta(days=1)
    else:
        last_day = pd.Timestamp(date(anio_fin, mes_fin + 1, 1)) - pd.Timedelta(days=1)
    end = pd.Timestamp(last_day.date()) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    return start, end


# ============================================================
# SECCIÓN 3 — GUI DE PROGRESO
# ============================================================

class ProgressUI:
    def __init__(self, root: tk.Tk,
                 title: str = "Calidad Participantes/Servicios - Progreso",
                 size:  str = "900x560"):
        self.root = root
        self.root.title(title)
        self.root.geometry(size)
        self.q: "queue.Queue[dict]" = queue.Queue()
        self.status_var = tk.StringVar(value="Listo para iniciar...")
        ttk.Label(self.root, textvariable=self.status_var,
                  wraplength=860).pack(padx=12, pady=(12, 6), anchor="w")
        self.pb = ttk.Progressbar(self.root, mode="determinate", maximum=100)
        self.pb.pack(padx=12, pady=(0, 10), fill="x")
        self.txt = tk.Text(self.root, height=22, wrap="word")
        self.txt.pack(padx=12, pady=(0, 12), fill="both", expand=True)
        self.txt.configure(state="disabled")
        self._indeterminate = False

    def _log(self, s: str) -> None:
        self.txt.configure(state="normal")
        self.txt.insert("end", s + "\n")
        self.txt.see("end")
        self.txt.configure(state="disabled")

    def _set_indeterminate(self, on: bool) -> None:
        if on and not self._indeterminate:
            self.pb.config(mode="indeterminate"); self.pb.start(10)
            self._indeterminate = True
        elif (not on) and self._indeterminate:
            self.pb.stop(); self.pb.config(mode="determinate")
            self._indeterminate = False

    def progress(self, stage: str, current, total, message: str) -> None:
        self.q.put({"type": "progress", "stage": stage,
                    "current": current, "total": total, "message": message})

    def done(self, output_path: str, meta_msg: str) -> None:
        self.q.put({"type": "done", "output": output_path, "meta": meta_msg})

    def error(self, err: str) -> None:
        self.q.put({"type": "error", "error": err})

    def poll(self) -> None:
        try:
            while True:
                msg = self.q.get_nowait()
                if msg["type"] == "progress":
                    self.status_var.set(msg["message"]); self._log(msg["message"])
                    cur, tot = msg.get("current"), msg.get("total")
                    if tot is None or tot <= 0 or cur is None:
                        self._set_indeterminate(True)
                    else:
                        self._set_indeterminate(False)
                        self.pb["value"] = int(max(0, min(100, cur / tot * 100)))
                elif msg["type"] == "done":
                    self._set_indeterminate(False); self.pb["value"] = 100
                    self.status_var.set("✅ Finalizado.")
                    self._log(f"✅ Archivo generado: {msg['output']}")
                    if msg.get("meta"):
                        self._log(msg["meta"])
                    messagebox.showinfo("Calidad Participantes/Servicios",
                                        f"Se generó:\n{msg['output']}")
                elif msg["type"] == "error":
                    self._set_indeterminate(False); self.status_var.set("❌ Error.")
                    self._log("❌ ERROR: " + msg["error"])
                    messagebox.showerror("Error", msg["error"])
        except queue.Empty:
            pass
        self.root.after(120, self.poll)

    def run(self) -> None:
        self.poll(); self.root.mainloop()


# ============================================================
# SECCIÓN 4 — CARGA Y NORMALIZACIÓN
# ============================================================

@dataclass
class LoadResult:
    input_path:    Path
    participantes: pd.DataFrame
    servicios:     pd.DataFrame


def parse_date_val(x) -> pd.Timestamp:
    if pd.isna(x): return pd.NaT
    if isinstance(x, (pd.Timestamp, datetime, date)):
        return pd.to_datetime(x, errors="coerce")
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        xx = float(x)
        if 10000 < xx < 60000:
            return pd.to_datetime(xx, unit="D", origin="1899-12-30", errors="coerce")
        return pd.NaT
    s = str(x).strip().replace("00:00:00", "").strip()
    if not s or s.lower() in ("nan", "none", "nat"): return pd.NaT
    if "/" in s: return pd.to_datetime(s, dayfirst=True, errors="coerce")
    if re.match(r"^\d{4}-\d{2}-\d{2}", s): return pd.to_datetime(s, errors="coerce")
    return pd.to_datetime(s, errors="coerce")


def _pick_sheet(dfs: Dict, names: List[str]) -> Optional[pd.DataFrame]:
    lk = {k.lower(): k for k in dfs}
    for n in names:
        if n.lower() in lk: return dfs[lk[n.lower()]]
    return None


def load_file(input_path: str, progress_cb) -> LoadResult:
    path = Path(input_path).expanduser().resolve()
    progress_cb("carga", None, None, "Cargando Excel (todas las hojas)...")
    if path.suffix.lower() == ".xlsx":
        dfs = pd.read_excel(path, sheet_name=None, engine="openpyxl", dtype=object)
    elif path.suffix.lower() == ".xls":
        dfs = pd.read_excel(path, sheet_name=None, engine="xlrd", dtype=object)
    else:
        raise RuntimeError("Formato no soportado. Use .xlsx o .xls.")
    progress_cb("carga", 1, 1, f"Excel cargado — {len(dfs)} hojas.")
    total = max(1, len(dfs))
    for i, sh in enumerate(list(dfs.keys()), 1):
        df = dfs[sh]
        df.columns = [norm_col(c) for c in df.columns]
        df = df.replace(r"^\s*$", np.nan, regex=True)
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all").dropna(axis=1, how="all")
        dfs[sh] = df
        progress_cb("prepro", i, total, f"Normalizada {i}/{total}: {sh}")
    part = _pick_sheet(dfs, ["participantes", "participants"])
    svc  = _pick_sheet(dfs, ["servicios", "services"])
    if part is None: raise RuntimeError(f"No encontré 'participantes'. Hojas: {list(dfs)}")
    if svc  is None: raise RuntimeError(f"No encontré 'servicios'. Hojas: {list(dfs)}")
    progress_cb("base", 1, 1,
                f"Hojas listas — participantes: {len(part)}, servicios: {len(svc)}.")
    return LoadResult(input_path=path, participantes=part, servicios=svc)


# ============================================================
# SECCIÓN 5 — UTILIDADES DE TEXTO Y COMPLETITUD
# ============================================================

def norm_col(c: str) -> str:
    s = str(c)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s).strip().upper()


def norm_text(s: str) -> str:
    return norm_col(s)


def is_filled(series: pd.Series) -> pd.Series:
    filled = ~series.isna()
    as_str = series.astype(str).str.strip()
    filled &= ~as_str.eq("")
    filled &= ~as_str.str.lower().isin(["nan", "none", "nat"])
    return filled


def find_first_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if norm_col(c) in df.columns: return norm_col(c)
    return None


def normalize_cat_series(s: pd.Series, fill: str) -> pd.Series:
    out = s.astype(str).str.strip()
    out = out.replace({"": np.nan, "NAN": np.nan, "NONE": np.nan, "NAT": np.nan})
    return out.fillna(fill)


def _safe_pct(num: float, den: float) -> float:
    return float("nan") if den == 0 else round(num / den * 100, 2)


# ============================================================
# SECCIÓN 6 — MAPAS DE CAMPOS MÍNIMOS Y CONSTANTES
# ============================================================

PARD_ONLY_FIELDS: List[str] = ["DEFENSOR/A", "FECHA APERTURA PARD"]
PARD_TIPO_KEYWORD = "PARD"

GRUPO_ETARIO_RANGOS: Dict[str, Tuple[int, int]] = {
    "0-5":   (0,   5),
    "6-12":  (6,  12),
    "13-17": (13, 17),
    "18-24": (18, 24),
    "24+":   (24, 150),
}

DOC_REGLAS: Dict[str, Dict] = {
    "TARJETA DE IDENTIDAD":  {"min": 10, "max": 11, "solo_numeros": True},
    "CEDULA DE CIUDADANIA":  {"min": 5,  "max": 10, "solo_numeros": True},
    "REGISTRO CIVIL":        {"min": 10, "max": 11, "solo_numeros": True},
    "CEDULA DE EXTRANJERIA": {"min": 4,  "max": 10, "solo_numeros": False},
    "PPT":                   {"min": 7,  "max": 15, "solo_numeros": False},
    "PEP":                   {"min": 7,  "max": 15, "solo_numeros": False},
    "PASAPORTE":             {"min": 5,  "max": 20, "solo_numeros": False},
    "DNI":                   {"min": 6,  "max": 12, "solo_numeros": False},
    "CEDULA VENEZOLANA":     {"min": 6,  "max": 10, "solo_numeros": False},
}
DOC_SKIP: Set[str] = {
    norm_text(t) for t in
    ["OTRO", "ACTA", "ACTA NACIMIENTO", "NACIDO VIVO", "RUMV", "SD", "MS",
     "SIN DOCUMENTO", "MS."]
}

SERVICIO_FAMILIA_REGLAS: Dict[str, Dict] = {
    "DFE": {"prohibe_acogida": True,  "origen_solo_advertencia": False},
    "FFC": {"prohibe_acogida": False, "origen_solo_advertencia": True},
    "FLC": {"prohibe_acogida": False, "origen_solo_advertencia": True},
    "SIL": {"prohibe_acogida": False, "origen_solo_advertencia": True},
}

TIPO_ESPERADO_POR_GRUPO: Dict[str, Set[str]] = {
    norm_text("CA_Cuidado_Alternativo"): {
        norm_text("NNAJ"), norm_text("NNAJ CON PARD"),
    },
    norm_text("FS_Fortalecimiento_Familiar"): {
        norm_text("C.P. - CUIDADOR/A PRINCIPAL"),
        norm_text("C.S. - CUIDADOR/A SECUNDARIO/A"),
    },
}


def build_required_maps() -> Tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    req_part: Dict[str, List[str]] = {
        "TIPO PARTICIPANTE": ["TIPO PARTICIPANTE"],
        "NOMBRES Y APELLIDOS DEL PARTICIPANTE": [
            "NOMBRES Y APELLIDOS DEL PARTICIPANTE",
            "NOMBRE COMPLETO", "NOMBRES Y APELLIDOS"],
        "TIPO DE DOCUMENTO":   ["TIPO DE DOCUMENTO"],
        "NUMERO DE DOCUMENTO": [
            "NUMERO DE DOCUMENTO", "NÚMERO DE DOCUMENTO",
            "DOCUMENTO", "NUM DOCUMENTO"],
        "FECHA DE NACIMIENTO":    ["FECHA DE NACIMIENTO"],
        "SEXO DEL PARTICIPANTE":  ["SEXO DEL PARTICIPANTE", "SEXO"],
        "GENERO DEL PARTICIPANTE": [
            "GENERO DEL PARTICIPANTE", "GÉNERO DEL PARTICIPANTE",
            "GENERO", "GÉNERO"],
        "DEPARTAMENTO (RESIDENCIA)": [
            "DEPARTAMENTO (RESIDENCIA)", "DEPARTAMENTO RESIDENCIA", "DEPARTAMENTO"],
        "MUNICIPIO (RESIDENCIA)": [
            "MUNICIPIO (RESIDENCIA)", "MUNICIPIO RESIDENCIA", "MUNICIPIO"],
        # Condicionales PARD
        "DEFENSOR/A":          ["DEFENSOR/A", "DEFENSOR", "DEFENSORA", "DEFENSOR(A)"],
        "FECHA APERTURA PARD": ["FECHA APERTURA PARD", "FECHA DE APERTURA PARD"],
    }
    req_svc: Dict[str, List[str]] = {
        "PROGRAMA": ["PROGRAMA"],
        "GRUPO DE SERVICIO (FEDERATIVOS)": [
            "GRUPO DE SERVICIO (FEDERATIVOS)", "GRUPO DE SERVICIO FEDERATIVOS",
            "GRUPO DE SERVICIO"],
        "SERVICIO (FEDERATIVOS)": [
            "SERVICIO (FEDERATIVOS)", "SERVICIO FEDERATIVOS", "SERVICIO"],
        "SUB-SERVICIO": ["SUB-SERVICIO", "SUB SERVICIO", "SUBSERVICIO"],
        "FAMILIA DE ORIGEN": [
            "FAMILIA DE ORIGEN", "ID FAMILIA DE ORIGEN EN DFE",
            "ID FAMILIA DE ORIGEN"],
        "FAMILIA / CASA DE ACOGIDA": [
            "FAMILIA / CASA DE ACOGIDA", "ID FAMILIA / CASA DE ACOGIDA",
            "ID FAMILIA CASA DE ACOGIDA"],
        "FECHA DE INGRESO A ALDEAS":    ["FECHA DE INGRESO A ALDEAS"],
        "FECHA DE ENTRADA AL SERVICIO": ["FECHA DE ENTRADA AL SERVICIO"],
        "MOTIVO DE ENTRADA AL SERVICIO": [
            "MOTIVO DE ENTRADA AL SERVICIO", "MOTIVO DE INGRESO AL SERVICIO"],
    }
    return req_part, req_svc


def default_output_path(input_path: Path, period: str) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return input_path.with_name(
        f"CALIDAD_PS_{period}_{input_path.stem}_{ts}.xlsx")


# ============================================================
# SECCIÓN 7 — COMPLETITUD (MÁSCARAS DE PRESENCIA)
# ============================================================

def compute_presence_masks(
    df: pd.DataFrame,
    required_map: Dict[str, List[str]],
) -> Tuple[Dict[str, Optional[str]], Dict[str, pd.Series], pd.Series]:
    col_map:   Dict[str, Optional[str]] = {}
    miss_mask: Dict[str, pd.Series]     = {}
    for logical, cands in required_map.items():
        col = find_first_col(df, cands)
        col_map[logical] = col
        miss_mask[logical] = (
            pd.Series([True] * len(df), index=df.index) if col is None
            else ~is_filled(df[col])
        )
    ok = pd.Series([True] * len(df), index=df.index)
    for f in required_map:
        ok &= ~miss_mask[f]
    return col_map, miss_mask, ok


def apply_pard_condition(
    miss_mask: Dict[str, pd.Series],
    pard_fields: List[str],
    is_pard_row: pd.Series,
) -> None:
    """Des-penaliza campos PARD en filas que NO son tipo PARD (in-place)."""
    for f in pard_fields:
        if f in miss_mask:
            miss_mask[f] = miss_mask[f] & is_pard_row.fillna(False)


def recompute_ok(
    miss_mask: Dict[str, pd.Series],
    fields: List[str],
    index: pd.Index,
) -> pd.Series:
    ok = pd.Series([True] * len(index), index=index)
    for f in fields:
        ok &= ~miss_mask[f]
    return ok


def missing_list_per_row(
    miss_mask: Dict[str, pd.Series], order: List[str]
) -> pd.Series:
    idx = next(iter(miss_mask.values())).index if miss_mask else pd.Index([])
    out = []
    for i in range(len(idx)):
        miss = [k for k in order if bool(miss_mask[k].iat[i])]
        out.append("; ".join(miss))
    return pd.Series(out, index=idx)


def n_missing_cells(
    miss_mask: Dict[str, pd.Series],
    idx: pd.Index,
    fields: List[str],
) -> int:
    return sum(int(miss_mask[f].loc[idx].sum()) for f in fields)


# ============================================================
# SECCIÓN 8 — VALIDACIONES LÓGICAS
# ============================================================

# ---- 8.1 Participantes ----

def check_fechas_nacimiento(df, col_fnac, ref_date):
    errores = []
    hoy = pd.Timestamp(ref_date)
    limite = pd.Timestamp("1920-01-01")
    for idx, val in df[col_fnac].items():
        dt = parse_date_val(val)
        if pd.isna(dt): continue
        if dt > hoy:
            errores.append((idx, f"Fecha nacimiento en el futuro: {dt.date()}"))
        elif dt < limite:
            errores.append((idx, f"Fecha nacimiento anterior a 1920: {dt.date()}"))
    return errores


def check_edad_vs_fnac(df, col_edad, col_fnac, ref_date):
    errores = []
    for idx, row in df.iterrows():
        fnac = parse_date_val(row[col_fnac])
        if pd.isna(fnac): continue
        try:
            edad_decl = int(float(str(row[col_edad])))
        except Exception:
            continue
        edad_calc = (ref_date - fnac.date()).days // 365
        if abs(edad_decl - edad_calc) > 1:
            errores.append((idx,
                f"Edad declarada {edad_decl} no coincide con nacimiento "
                f"{fnac.date()} (calculada: {edad_calc})"))
    return errores


def check_grupo_etario_vs_edad(df, col_grupo, col_edad):
    errores = []
    rangos_norm = {norm_text(k): v for k, v in GRUPO_ETARIO_RANGOS.items()}
    for idx, row in df.iterrows():
        try:
            edad = int(float(str(row[col_edad])))
        except Exception:
            continue
        grupo_raw  = str(row[col_grupo])
        grupo_norm = norm_text(grupo_raw)
        if grupo_norm in (norm_text("Sin información"), "NAN", "NONE", ""):
            continue
        encontrado = False
        for clave, (mn, mx) in rangos_norm.items():
            if clave in grupo_norm:
                encontrado = True
                if not (mn <= edad <= mx):
                    errores.append((idx,
                        f"Grupo etario '{grupo_raw}' (rango {mn}-{mx}) "
                        f"no corresponde a edad {edad}"))
                break
        if not encontrado:
            errores.append((idx,
                f"Grupo etario no reconocido: '{grupo_raw}'"))
    return errores


def check_tipo_doc_vs_numero(df, col_tipo, col_num):
    errores, advertencias = [], []
    reglas_norm = {norm_text(k): v for k, v in DOC_REGLAS.items()}
    for idx, row in df.iterrows():
        tipo_raw  = str(row[col_tipo]).strip()
        tipo_norm = norm_text(tipo_raw)
        if tipo_norm in ("", "NAN", "NONE"): continue
        if tipo_norm in DOC_SKIP: continue
        num_raw  = str(row[col_num]).strip() if col_num and col_num in df.columns else ""
        num_norm = norm_text(num_raw)
        num_tiene = num_norm not in ("", "NAN", "NONE")
        regla = next((r for k, r in reglas_norm.items() if k in tipo_norm), None)
        if regla is None:
            advertencias.append((idx,
                f"Tipo documento no reconocido: '{tipo_raw}'"))
            continue
        if not num_tiene:
            continue  # completitud ya lo informa
        num_limpio = re.sub(r"\s+", "", num_raw)
        if regla["solo_numeros"] and not num_limpio.isdigit():
            errores.append((idx,
                f"Documento '{tipo_raw}': número '{num_limpio}' "
                f"debe contener solo dígitos"))
        if not (regla["min"] <= len(num_limpio) <= regla["max"]):
            errores.append((idx,
                f"Documento '{tipo_raw}': longitud {len(num_limpio)} "
                f"fuera del rango [{regla['min']}-{regla['max']}]"))
    return errores, advertencias


def compute_duplicate_participante_ids(
    participantes, pid_col, col_tipo, col_num, col_nombre_norm
) -> Tuple[Set[str], Set[str]]:
    """
    Duplicados por (tipo+número); fallback a nombre normalizado cuando
    el número está vacío.
    """
    pids = participantes[pid_col].astype(str).str.strip()
    dup_doc_pids: Set[str]    = set()
    dup_nombre_pids: Set[str] = set()
    if col_tipo and col_num:
        tipos      = participantes[col_tipo].astype(str).str.strip().map(norm_text)
        nums       = participantes[col_num].astype(str).str.strip()
        num_filled = is_filled(participantes[col_num])
        keys_doc   = (tipos + "||" + nums).where(num_filled, np.nan)
        dup_doc_pids = set(pids[keys_doc.duplicated(keep=False) & num_filled].tolist())
        if col_nombre_norm:
            sin_num = ~num_filled
            nombres = participantes[col_nombre_norm].astype(str).str.strip().map(norm_text)
            dup_nombre_pids = set(pids[nombres.duplicated(keep=False) & sin_num].tolist())
    return dup_doc_pids, dup_nombre_pids


def check_fecha_apertura_pard(df, col_fapert, col_fnac, is_pard, ref_date):
    errores = []
    hoy = pd.Timestamp(ref_date)
    for idx, row in df[is_pard.fillna(False)].iterrows():
        fapert = parse_date_val(row[col_fapert])
        if pd.isna(fapert): continue
        if fapert > hoy:
            errores.append((idx, f"Fecha apertura PARD en el futuro: {fapert.date()}"))
        if col_fnac and col_fnac in df.columns:
            fnac = parse_date_val(row[col_fnac])
            if not pd.isna(fnac) and fapert < fnac:
                errores.append((idx,
                    f"Fecha apertura PARD ({fapert.date()}) anterior a "
                    f"nacimiento ({fnac.date()})"))
    return errores


# ---- 8.2 Servicios ----

def check_fechas_servicio(df, col_entrada, col_aldeas, col_salida, ref_date):
    errores = []
    hoy = pd.Timestamp(ref_date)
    for idx, row in df.iterrows():
        entrada = parse_date_val(row[col_entrada])
        if pd.isna(entrada): continue
        if entrada > hoy:
            errores.append((idx,
                f"FECHA ENTRADA SERVICIO en el futuro: {entrada.date()}"))
        if col_aldeas and col_aldeas in df.columns:
            aldeas = parse_date_val(row[col_aldeas])
            if not pd.isna(aldeas):
                if aldeas > hoy:
                    errores.append((idx,
                        f"FECHA INGRESO ALDEAS en el futuro: {aldeas.date()}"))
                if aldeas > entrada:
                    errores.append((idx,
                        f"FECHA INGRESO ALDEAS ({aldeas.date()}) "
                        f"posterior a FECHA ENTRADA ({entrada.date()})"))
        if col_salida and col_salida in df.columns:
            salida = parse_date_val(row[col_salida])
            if not pd.isna(salida) and salida < entrada:
                errores.append((idx,
                    f"FECHA SALIDA ({salida.date()}) anterior a "
                    f"FECHA ENTRADA ({entrada.date()})"))
    return errores


def check_familia_vs_servicio(df, col_servicio, col_sub_servicio,
                               col_fam_orig, col_fam_acog, col_derechos):
    """
    Valida la relación entre familia y tipo de servicio.

    Reglas:
    · DFE  con FAMILIA/CASA DE ACOGIDA diligenciada → error crítico.
    · FFC, FLC, SIL-Comunidad Juvenil con FAMILIA DE ORIGEN vacía → advertencia,
      EXCEPTO si DERECHOS O CAPACIDADES = ADOPTABILIDAD (esos NNAJ no regresarán
      con su familia de origen, la advertencia no aplica).
    """
    errores, advertencias = [], []

    def _adoptabilidad(row):
        if not col_derechos or col_derechos not in df.columns:
            return False
        return norm_text(str(row[col_derechos])) == norm_text("ADOPTABILIDAD")

    def _es_sil_cj(row):
        """SIL con sub-servicio Comunidad Juvenil."""
        if not col_sub_servicio or col_sub_servicio not in df.columns:
            return False
        return norm_text(str(row[col_sub_servicio])) == norm_text("COMUNIDAD JUVENIL")

    for idx, row in df.iterrows():
        serv_norm = norm_text(str(row[col_servicio]))
        tipo_svc  = next(
            (k for k in SERVICIO_FAMILIA_REGLAS if k in serv_norm), None)
        if tipo_svc is None: continue
        regla = SERVICIO_FAMILIA_REGLAS[tipo_svc]

        # DFE: no debe tener familia de acogida diligenciada
        if regla["prohibe_acogida"] and col_fam_acog and col_fam_acog in df.columns:
            val = row[col_fam_acog]
            if is_filled(pd.Series([val])).iloc[0]:
                errores.append((idx,
                    f"ALERTA CRÍTICA: servicio DFE tiene FAMILIA/CASA DE ACOGIDA "
                    f"'{val}'. DFE trabaja con familia de ORIGEN, no de acogida. "
                    f"Verificar asignación incorrecta de participante o familia."))

        # FFC/FLC: advertencia si FAMILIA DE ORIGEN vacía y no es adoptabilidad
        if regla["origen_solo_advertencia"] and col_fam_orig and col_fam_orig in df.columns:
            if not is_filled(pd.Series([row[col_fam_orig]])).iloc[0]:
                if not _adoptabilidad(row):
                    advertencias.append((idx,
                        f"FAMILIA DE ORIGEN vacía en servicio {tipo_svc} "
                        f"— se espera en los primeros 3 meses de atención"))

        # SIL-Comunidad Juvenil: advertencia si FAMILIA DE ORIGEN vacía y no es adoptabilidad
        if tipo_svc == "SIL" and _es_sil_cj(row):
            if col_fam_orig and col_fam_orig in df.columns:
                if not is_filled(pd.Series([row[col_fam_orig]])).iloc[0]:
                    if not _adoptabilidad(row):
                        advertencias.append((idx,
                            f"FAMILIA DE ORIGEN vacía en servicio SIL-Comunidad Juvenil "
                            f"— se espera en los primeros 3 meses de atención"))

    return errores, advertencias


def compute_duplicate_servicio_ids(servicios, sid_col) -> Set[str]:
    sids = servicios[sid_col].astype(str).str.strip()
    dup  = sids.duplicated(keep=False) & is_filled(servicios[sid_col])
    return set(sids[dup].tolist())


# ---- 8.3 Cruzados ----

def check_fecha_entrada_vs_nacimiento(df, col_entrada, col_fnac):
    errores = []
    for idx, row in df.iterrows():
        entrada = parse_date_val(row[col_entrada])
        fnac    = parse_date_val(row[col_fnac])
        if pd.isna(entrada) or pd.isna(fnac): continue
        if entrada < fnac:
            errores.append((idx,
                f"FECHA ENTRADA SERVICIO ({entrada.date()}) "
                f"anterior a FECHA NACIMIENTO ({fnac.date()})"))
    return errores


def check_tipo_part_vs_grupo_svc(df, col_tipo_part, col_grupo_svc):
    advertencias = []
    for idx, row in df.iterrows():
        grupo_norm = norm_text(str(row[col_grupo_svc]))
        tipo_norm  = norm_text(str(row[col_tipo_part]))
        if not grupo_norm or tipo_norm in ("", "NAN", "NONE"): continue
        esperados = TIPO_ESPERADO_POR_GRUPO.get(grupo_norm)
        if esperados and tipo_norm not in esperados:
            advertencias.append((idx,
                f"Tipo participante '{row[col_tipo_part]}' inusual "
                f"para grupo '{row[col_grupo_svc]}'"))
    return advertencias


# ---- 8.4 Orquestador ----

def run_all_logic_checks(
    merged, participantes, servicios,
    part_colmap, svc_colmap, aux_cols,
    is_pard, dup_doc_pids, dup_nombre_pids, dup_svc_ids,
    pid_part_col, sid_svc_col, ref_date, progress_cb,
) -> Tuple[Dict[int, List[str]], Dict[int, List[str]]]:
    errores:      Dict[int, List[str]] = {}
    advertencias: Dict[int, List[str]] = {}

    def _ae(lst):
        for i, m in lst: errores.setdefault(i, []).append(m)

    def _aa(lst):
        for i, m in lst: advertencias.setdefault(i, []).append(m)

    col_fnac      = aux_cols.get("FECHA DE NACIMIENTO")
    col_edad      = aux_cols.get("EDAD")
    col_grupo_et  = aux_cols.get("GRUPO ETARIO")
    col_tipo_doc  = part_colmap.get("TIPO DE DOCUMENTO")
    col_num_doc   = part_colmap.get("NUMERO DE DOCUMENTO")
    col_fapert    = part_colmap.get("FECHA APERTURA PARD")
    col_tipo_part = aux_cols.get("TIPO PARTICIPANTE")
    col_entrada   = svc_colmap.get("FECHA DE ENTRADA AL SERVICIO")
    col_aldeas    = svc_colmap.get("FECHA DE INGRESO A ALDEAS")
    col_salida    = find_first_col(merged, ["FECHA DE SALIDA DEL SERVICIO"])
    col_servicio  = svc_colmap.get("SERVICIO (FEDERATIVOS)")
    col_sub_svc   = svc_colmap.get("SUB-SERVICIO")
    col_grupo_svc = svc_colmap.get("GRUPO DE SERVICIO (FEDERATIVOS)")
    col_fam_orig  = svc_colmap.get("FAMILIA DE ORIGEN")
    col_fam_acog  = svc_colmap.get("FAMILIA / CASA DE ACOGIDA")
    col_derechos  = find_first_col(merged, [
        "DERECHOS O CAPACIDADES (CALCULO 18 MESES)",
        "DERECHOS O CAPACIDADES"])

    progress_cb("logic", None, None, "Lógica: fechas de nacimiento...")
    if col_fnac:
        _ae(check_fechas_nacimiento(merged, col_fnac, ref_date))

    progress_cb("logic", None, None, "Lógica: edad vs. nacimiento...")
    if col_edad and col_fnac:
        _ae(check_edad_vs_fnac(merged, col_edad, col_fnac, ref_date))

    progress_cb("logic", None, None, "Lógica: grupo etario vs. edad...")
    if col_grupo_et and col_edad:
        _ae(check_grupo_etario_vs_edad(merged, col_grupo_et, col_edad))

    progress_cb("logic", None, None, "Lógica: tipo documento vs. número...")
    if col_tipo_doc:
        e, a = check_tipo_doc_vs_numero(merged, col_tipo_doc, col_num_doc)
        _ae(e); _aa(a)

    progress_cb("logic", None, None, "Lógica: duplicados de participante...")
    for idx, row in merged.iterrows():
        pid = str(row.get("_PID_KEY", ""))
        if pid in dup_doc_pids:
            errores.setdefault(idx, []).append(
                f"Número de documento duplicado en la tabla (PID: {pid})")
        elif pid in dup_nombre_pids:
            advertencias.setdefault(idx, []).append(
                f"Sin documento — nombre normalizado duplicado, posible registro doble (PID: {pid})")

    progress_cb("logic", None, None, "Lógica: duplicados de ID servicio...")
    if sid_svc_col in merged.columns:
        for idx, row in merged.iterrows():
            sid = str(row[sid_svc_col]).strip()
            if sid in dup_svc_ids:
                errores.setdefault(idx, []).append(
                    f"ID SERVICIO '{sid}' duplicado — rompe integridad relacional")

    progress_cb("logic", None, None, "Lógica: fecha apertura PARD...")
    if col_fapert:
        _ae(check_fecha_apertura_pard(merged, col_fapert, col_fnac, is_pard, ref_date))

    progress_cb("logic", None, None, "Lógica: fechas de servicio...")
    if col_entrada:
        _ae(check_fechas_servicio(merged, col_entrada, col_aldeas, col_salida, ref_date))

    progress_cb("logic", None, None, "Lógica: familia vs. tipo de servicio...")
    if col_servicio:
        e, a = check_familia_vs_servicio(
            merged, col_servicio, col_sub_svc,
            col_fam_orig, col_fam_acog, col_derechos)
        _ae(e); _aa(a)

    progress_cb("logic", None, None, "Lógica: servicios huérfanos...")
    part_ids_set = set(participantes[pid_part_col].astype(str).str.strip().dropna())
    for idx, row in merged.iterrows():
        pid = str(row.get("_PID_KEY", "")).strip()
        if pid and pid.lower() not in ("nan", "none", "nat", "") \
                and pid not in part_ids_set:
            errores.setdefault(idx, []).append(
                f"SERVICIO HUÉRFANO: PID '{pid}' no existe en participantes "
                f"— rompe modelo relacional")

    progress_cb("logic", None, None, "Lógica: fecha entrada vs. nacimiento...")
    if col_entrada and col_fnac:
        _ae(check_fecha_entrada_vs_nacimiento(merged, col_entrada, col_fnac))

    progress_cb("logic", None, None, "Lógica: tipo participante vs. grupo servicio...")
    if col_tipo_part and col_grupo_svc:
        _aa(check_tipo_part_vs_grupo_svc(merged, col_tipo_part, col_grupo_svc))

    n_err = sum(len(v) for v in errores.values())
    n_adv = sum(len(v) for v in advertencias.values())
    progress_cb("logic", 1, 1,
                f"Validaciones lógicas: {n_err} errores | {n_adv} advertencias")
    return errores, advertencias


# ============================================================
# SECCIÓN 9 — PIPELINE PRINCIPAL + RESÚMENES
# ============================================================

def run_checker(
    input_path: Path,
    progress_cb,
    mes_ini: int, anio_ini: int,
    mes_fin: int, anio_fin: int,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    """
    Orquesta carga → cohorte → completitud → lógica → resúmenes.
    Retorna (resumen_fila, resumen_celda, casos, meta).
    Estructura idéntica a comprobador_estadisticas_2025.run_checker().
    """
    lr = load_file(str(input_path), progress_cb)
    participantes = lr.participantes
    servicios     = lr.servicios

    req_part, req_svc = build_required_maps()
    part_fields = list(req_part.keys())
    svc_fields  = list(req_svc.keys())
    all_fields  = part_fields + svc_fields

    periodo_str = period_display(mes_ini, anio_ini, mes_fin, anio_fin)
    ref_date    = date.today()

    # --- Columnas clave ---
    progress_cb("setup", None, None, "Identificando columnas clave...")
    pid_part    = find_first_col(participantes, [
        "ID DEL PARTICIPANTE (PRIMARIA)", "ID DEL PARTICIPANTE", "ID PARTICIPANTE"])
    pid_svc     = find_first_col(servicios, ["ID DEL PARTICIPANTE", "ID PARTICIPANTE"])
    sid_svc     = find_first_col(servicios, ["ID SERVICIO", "ID DEL SERVICIO"])
    prog_col    = find_first_col(servicios, ["PROGRAMA"])
    entrada_col = find_first_col(servicios, ["FECHA DE ENTRADA AL SERVICIO"])
    tipo_part_col = find_first_col(participantes, ["TIPO PARTICIPANTE"])

    for col, nombre in [
        (pid_part,    "ID participante (hoja participantes)"),
        (pid_svc,     "ID participante (hoja servicios)"),
        (sid_svc,     "ID SERVICIO"),
        (prog_col,    "PROGRAMA"),
        (entrada_col, "FECHA DE ENTRADA AL SERVICIO"),
    ]:
        if col is None:
            raise RuntimeError(f"Columna requerida no encontrada: {nombre}")

    # --- Cohorte del período ---
    start_p, end_p = period_bounds(mes_ini, anio_ini, mes_fin, anio_fin)
    progress_cb("cohorte", None, None,
                f"Filtrando cohorte {periodo_str}: {start_p.date()} → {end_p.date()}...")

    svc = servicios.copy()
    svc["_ENTRADA_DT"] = svc[entrada_col].map(parse_date_val)
    svc_cohort = svc.loc[
        svc["_ENTRADA_DT"].notna() &
        (svc["_ENTRADA_DT"] >= start_p) &
        (svc["_ENTRADA_DT"] <= end_p)
    ].copy()
    svc_cohort["_ROW_ORIG"] = svc_cohort.index
    svc_cohort = svc_cohort.reset_index(drop=True)

    n_archivo  = len(servicios)
    n_periodo  = len(svc_cohort)
    n_fuera    = n_archivo - n_periodo

    progress_cb("cohorte", 1, 1,
                f"Cohorte: {n_periodo} servicios "
                f"(de {n_archivo} en el archivo — {n_fuera} fuera del período).")

    if n_periodo == 0:
        raise RuntimeError(
            f"No hay servicios para el período {periodo_str}. "
            f"Verifica que el archivo contenga datos de ese período.")

    # --- Precomputar duplicados y huérfanos ---
    progress_cb("precomp", None, None, "Precomputando duplicados y huérfanos...")
    col_tipo_doc_p   = find_first_col(participantes, req_part["TIPO DE DOCUMENTO"])
    col_num_doc_p    = find_first_col(participantes, req_part["NUMERO DE DOCUMENTO"])
    col_nombre_p     = find_first_col(participantes, [
        "NOMBRES Y APELLIDOS DEL PARTICIPANTE", "NOMBRE COMPLETO"])

    dup_doc_pids, dup_nombre_pids = compute_duplicate_participante_ids(
        participantes, pid_part, col_tipo_doc_p, col_num_doc_p, col_nombre_p)
    dup_svc_ids = compute_duplicate_servicio_ids(servicios, sid_svc)

    pids_con_svc  = set(servicios[pid_svc].astype(str).str.strip().dropna())
    pids_part_all = participantes[pid_part].astype(str).str.strip()
    orphan_mask   = ~pids_part_all.isin(pids_con_svc) & is_filled(participantes[pid_part])
    orphans_part  = participantes[orphan_mask].copy().reset_index(drop=True)

    progress_cb("precomp", 1, 1,
                f"Dup. doc: {len(dup_doc_pids)} | Dup. nombre: {len(dup_nombre_pids)} | "
                f"Dup. svc: {len(dup_svc_ids)} | "
                f"Participantes huérfanos: {len(orphans_part)}")

    # --- Completitud SERVICIOS ---
    progress_cb("eval", None, None, "Completitud: SERVICIOS...")
    svc_colmap, svc_missmask, svc_ok = compute_presence_masks(svc_cohort, req_svc)

    # Ajustes de completitud por tipo de servicio:
    #   FAMILIA DE ORIGEN         → no requerida en FFC/FLC (solo advertencia)
    #   FAMILIA / CASA DE ACOGIDA → no requerida en:
    #       · DFE  (trabaja con familia de ORIGEN, nunca de acogida)
    #       · SIL con SUB-SERVICIO = "VIVIENDA CON ASESORAMIENTO"
    #         (el NNA vive de forma independiente, no en casa de acogida)
    col_serv_real = svc_colmap.get("SERVICIO (FEDERATIVOS)")
    col_sub_real  = svc_colmap.get("SUB-SERVICIO")
    if col_serv_real:
        serv_norm_c = svc_cohort[col_serv_real].astype(str).map(norm_text)
        is_ffc_flc  = serv_norm_c.str.contains(r"FFC|FLC", na=False)
        is_dfe      = serv_norm_c.str.contains(r"DFE",     na=False)
        is_sil      = serv_norm_c.str.contains(r"SIL",     na=False)
    else:
        is_ffc_flc = pd.Series(False, index=svc_cohort.index)
        is_dfe     = pd.Series(False, index=svc_cohort.index)
        is_sil     = pd.Series(False, index=svc_cohort.index)

    if col_sub_real:
        sub_norm_c = svc_cohort[col_sub_real].astype(str).map(norm_text)
        is_sil_vivienda = is_sil & sub_norm_c.str.contains("VIVIENDA CON ASESORAMIENTO", na=False)
    else:
        is_sil_vivienda = pd.Series(False, index=svc_cohort.index)

    if "FAMILIA DE ORIGEN" in svc_missmask:
        svc_missmask["FAMILIA DE ORIGEN"] = (
            svc_missmask["FAMILIA DE ORIGEN"] & ~is_ffc_flc)
    if "FAMILIA / CASA DE ACOGIDA" in svc_missmask:
        svc_missmask["FAMILIA / CASA DE ACOGIDA"] = (
            svc_missmask["FAMILIA / CASA DE ACOGIDA"] & ~is_dfe & ~is_sil_vivienda)
    svc_ok = recompute_ok(svc_missmask, svc_fields, svc_cohort.index)

    # --- Completitud PARTICIPANTES (regla PARD) ---
    progress_cb("eval", None, None, "Completitud: PARTICIPANTES (regla PARD)...")
    part_colmap, part_missmask, _ = compute_presence_masks(participantes, req_part)

    is_pard_part = (
        participantes[tipo_part_col].astype(str).str.upper()
        .str.contains(PARD_TIPO_KEYWORD, na=False)
        if tipo_part_col
        else pd.Series(False, index=participantes.index)
    )
    apply_pard_condition(part_missmask, PARD_ONLY_FIELDS, is_pard_part)
    part_ok = recompute_ok(part_missmask, part_fields, participantes.index)
    part_ids    = participantes[pid_part].astype(str).str.strip()
    part_ok_map = pd.Series(part_ok.values, index=part_ids).to_dict()

    # --- Merge cohorte ↔ participantes ---
    aux_col_names = {
        "FECHA DE NACIMIENTO": find_first_col(participantes, ["FECHA DE NACIMIENTO"]),
        "EDAD":                find_first_col(participantes, ["EDAD"]),
        "GRUPO ETARIO":        find_first_col(participantes, ["GRUPO ETARIO"]),
        "TIPO PARTICIPANTE":   tipo_part_col,
        "NOMBRE":              col_nombre_p,
    }
    part_keep = ([pid_part] +
                 [c for c in part_colmap.values() if c is not None] +
                 [c for c in aux_col_names.values() if c is not None])
    part_keep = list(dict.fromkeys(part_keep))
    part_min  = participantes[part_keep].copy()
    part_min["_PID_KEY"] = part_ids.values
    part_min["_ES_PARD"] = is_pard_part.values
    part_min  = part_min.drop_duplicates(subset="_PID_KEY")
    svc_cohort["_PID_KEY"] = svc_cohort[pid_svc].astype(str).str.strip()

    progress_cb("merge", None, None, "Uniendo cohorte ↔ participantes...")
    merged = svc_cohort.merge(
        part_min.drop(columns=[pid_part], errors="ignore"),
        on="_PID_KEY", how="left")

    participant_exists  = merged["_PID_KEY"].isin(set(part_ids.dropna()))
    ok_part_for_svc     = merged["_PID_KEY"].map(lambda x: bool(part_ok_map.get(x, False)))
    is_pard_merged      = pd.Series(
        merged.get("_ES_PARD", pd.Series(False, index=merged.index)).values,
        index=merged.index)

    # Máscaras de completitud de participante en merged
    part_missmask_on_svc: Dict[str, pd.Series] = {}
    for logical, actual in part_colmap.items():
        if actual is None or actual not in merged.columns:
            part_missmask_on_svc[logical] = pd.Series(True, index=merged.index)
        else:
            part_missmask_on_svc[logical] = ~is_filled(merged[actual])
    for logical in part_missmask_on_svc:
        part_missmask_on_svc[logical] |= ~participant_exists
    apply_pard_condition(part_missmask_on_svc, PARD_ONLY_FIELDS, is_pard_merged)

    part_miss_list = missing_list_per_row(part_missmask_on_svc, part_fields)
    part_miss_list = part_miss_list.where(
        participant_exists,
        "PARTICIPANTE_NO_ENCONTRADO; " + part_miss_list)
    svc_miss_list = missing_list_per_row(svc_missmask, svc_fields)

    # FALTANTES combinados (una sola columna para la pestaña casos)
    def _concat_miss(p, s):
        partes = [x for x in [p, s] if x and x != ""]
        return "; ".join(partes)

    faltantes = pd.Series(
        [_concat_miss(p, s)
         for p, s in zip(part_miss_list, svc_miss_list)],
        index=merged.index)

    ok_completitud = svc_ok & ok_part_for_svc

    # --- Validaciones lógicas ---
    aux_cols_merged = {
        k: (v if v and v in merged.columns else None)
        for k, v in aux_col_names.items()
    }
    aux_cols_merged["FECHA DE NACIMIENTO"] = (
        aux_col_names["FECHA DE NACIMIENTO"]
        if aux_col_names["FECHA DE NACIMIENTO"] and
           aux_col_names["FECHA DE NACIMIENTO"] in merged.columns
        else None
    )

    errores_dict, advertencias_dict = run_all_logic_checks(
        merged=merged, participantes=participantes, servicios=servicios,
        part_colmap=part_colmap, svc_colmap=svc_colmap,
        aux_cols=aux_cols_merged,
        is_pard=is_pard_merged,
        dup_doc_pids=dup_doc_pids, dup_nombre_pids=dup_nombre_pids,
        dup_svc_ids=dup_svc_ids,
        pid_part_col=pid_part, sid_svc_col=sid_svc,
        ref_date=ref_date, progress_cb=progress_cb,
    )

    errores_serie = pd.Series(
        {idx: "; ".join(msgs) for idx, msgs in errores_dict.items()}, dtype=str
    ).reindex(merged.index, fill_value="")
    advertencias_serie = pd.Series(
        {idx: "; ".join(msgs) for idx, msgs in advertencias_dict.items()}, dtype=str
    ).reindex(merged.index, fill_value="")

    ok_logica = errores_serie.eq("")
    ok_row    = ok_completitud & ok_logica
    n_err_ind = sum(len(v) for v in errores_dict.values())

    # --- Columnas de agrupación ---
    # PROGRAMA  → hoja servicios (columna PROGRAMA)
    # DEPARTAMENTO → DEPARTAMENTO (RESIDENCIA) de participantes, vinculado por ID del participante
    dep_actual = part_colmap.get("DEPARTAMENTO (RESIDENCIA)")
    mun_actual = part_colmap.get("MUNICIPIO (RESIDENCIA)")

    dep_series = (merged[dep_actual] if dep_actual and dep_actual in merged.columns
                  else pd.Series(np.nan, index=merged.index))
    dep  = normalize_cat_series(dep_series, "SIN_DEPARTAMENTO")
    prog = normalize_cat_series(merged[prog_col], "SIN_PROGRAMA")

    # --- RESÚMENES (estructura idéntica a estadísticas_2025) ---
    progress_cb("sum", None, None,
                "Construyendo resúmenes por PROGRAMA + DEPARTAMENTO...")
    grouped      = pd.DataFrame({"P": prog, "D": dep}).groupby(["P", "D"]).groups
    total_groups = max(1, len(grouped))
    rows_fila, rows_celda = [], []

    for i, ((pname, dname), idx) in enumerate(grouped.items(), 1):
        n        = len(idx)
        n_comp   = int(ok_completitud.loc[idx].sum())
        n_sin_log = int(ok_logica.loc[idx].sum())
        n_ok     = int(ok_row.loc[idx].sum())
        n_log_g  = n - n_sin_log

        rows_fila.append({
            "PROGRAMA":                    pname,
            "DEPARTAMENTO":                dname,
            "N_REGISTROS":                 n,
            "N_FILAS_COMPLETAS":           n_comp,
            "PCT_COMPLETITUD":             _safe_pct(n_comp,    n),
            "N_FILAS_SIN_ERRORES_LOGICOS": n_sin_log,
            "PCT_LOGICA_OK":               _safe_pct(n_sin_log, n),
            "N_FILAS_OK_TOTAL":            n_ok,
            "PCT_CALIDAD_TOTAL":           _safe_pct(n_ok,      n),
        })

        miss_p     = n_missing_cells(part_missmask_on_svc, idx, part_fields)
        miss_s     = n_missing_cells(svc_missmask,          idx, svc_fields)
        n_vacias   = miss_p + miss_s
        total_c    = n * len(all_fields)
        n_err_g    = sum(len(errores_dict[ii]) for ii in idx if ii in errores_dict)
        celdas_inv = n_vacias + n_log_g
        rows_celda.append({
            "PROGRAMA":                    pname,
            "DEPARTAMENTO":                dname,
            "N_REGISTROS":                 n,
            "TOTAL_CELDAS_ESPERADAS":      total_c,
            "N_CELDAS_VACIAS":             n_vacias,
            "PCT_CELDAS_DILIGENCIADAS":    _safe_pct(total_c - n_vacias, total_c),
            "N_FILAS_SIN_ERRORES_LOGICOS": n_sin_log,
            "PCT_LOGICA_OK":               _safe_pct(n_sin_log, n),
            "N_ERRORES_LOGICOS":           n_err_g,
            "PCT_CAMPOS_VALIDOS_TOTAL":    _safe_pct(total_c - celdas_inv, total_c),
        })
        progress_cb("sum", i, total_groups,
                    f"Resumen {i}/{total_groups}: {pname} | {dname}")

    n_tot    = n_periodo
    n_comp_t = int(ok_completitud.sum())
    n_sinl_t = int(ok_logica.sum())
    n_ok_t   = int(ok_row.sum())
    miss_p_t = n_missing_cells(part_missmask_on_svc, merged.index, part_fields)
    miss_s_t = n_missing_cells(svc_missmask,          merged.index, svc_fields)
    n_vac_t  = miss_p_t + miss_s_t
    tot_c_t  = n_tot * len(all_fields)
    celdas_inv_t = n_vac_t + (n_tot - n_sinl_t)

    rows_fila.append({
        "PROGRAMA": "TOTAL_GENERAL", "DEPARTAMENTO": "TODOS",
        "N_REGISTROS": n_tot,
        "N_FILAS_COMPLETAS":           n_comp_t,
        "PCT_COMPLETITUD":             _safe_pct(n_comp_t,    n_tot),
        "N_FILAS_SIN_ERRORES_LOGICOS": n_sinl_t,
        "PCT_LOGICA_OK":               _safe_pct(n_sinl_t,   n_tot),
        "N_FILAS_OK_TOTAL":            n_ok_t,
        "PCT_CALIDAD_TOTAL":           _safe_pct(n_ok_t,     n_tot),
    })
    rows_celda.append({
        "PROGRAMA":                    "TOTAL_GENERAL",
        "DEPARTAMENTO":                "TODOS",
        "N_REGISTROS":                 n_tot,
        "TOTAL_CELDAS_ESPERADAS":      tot_c_t,
        "N_CELDAS_VACIAS":             n_vac_t,
        "PCT_CELDAS_DILIGENCIADAS":    _safe_pct(tot_c_t - n_vac_t, tot_c_t),
        "N_FILAS_SIN_ERRORES_LOGICOS": n_sinl_t,
        "PCT_LOGICA_OK":               _safe_pct(n_sinl_t, n_tot),
        "N_ERRORES_LOGICOS":           n_err_ind,
        "PCT_CAMPOS_VALIDOS_TOTAL":    _safe_pct(tot_c_t - celdas_inv_t, tot_c_t),
    })

    resumen_fila  = pd.DataFrame(rows_fila).sort_values( ["PROGRAMA", "DEPARTAMENTO"])
    resumen_celda = pd.DataFrame(rows_celda).sort_values(["PROGRAMA", "DEPARTAMENTO"])

    # --- CASOS A CORREGIR (estructura idéntica a estadísticas_2025) ---
    progress_cb("casos", None, None, "Construyendo guía de casos a corregir...")

    col_nombre_merged = aux_cols_merged.get("NOMBRE")
    col_tipo_merged   = aux_cols_merged.get("TIPO PARTICIPANTE")
    col_tdoc_merged   = part_colmap.get("TIPO DE DOCUMENTO")
    col_ndoc_merged   = part_colmap.get("NUMERO DE DOCUMENTO")

    def _col(c):
        return (merged[c] if c and c in merged.columns
                else pd.Series([np.nan] * len(merged), index=merged.index))

    # PESTAÑA_ORIGEN: indica dónde ir a corregir en el Excel fuente
    def _origen(idx):
        fp = str(faltantes.at[idx])
        fe = str(errores_serie.at[idx])
        tiene_p = bool(fp and fp not in ("", "nan"))
        tiene_e = bool(fe and fe not in ("", "nan"))
        if tiene_p and tiene_e:  return "participantes y servicios"
        if tiene_p:
            # distinguir si el faltante es de participante, servicio o ambos
            fp_p = str(part_miss_list.at[idx])
            fp_s = str(svc_miss_list.at[idx])
            if fp_p and fp_s and fp_p not in ("", "nan") and fp_s not in ("", "nan"):
                return "participantes y servicios"
            if fp_p and fp_p not in ("", "nan"):
                return "participantes"
            return "servicios"
        if tiene_e:              return "lógica (revisar ambas hojas)"
        return ""

    pestana_origen = pd.Series(
        [_origen(i) for i in merged.index], index=merged.index)

    n_problemas = (
        faltantes.apply(       lambda s: 0 if not s else len(str(s).split("; "))) +
        errores_serie.apply(   lambda s: 0 if not s else len(str(s).split("; ")))
    )

    # Columnas: misma estructura que estadísticas_2025, con ID_SERVICIO añadido
    casos_df = pd.DataFrame({
        "PROGRAMA":            prog,
        "DEPARTAMENTO":        dep,
        "MUNICIPIO":           (merged[mun_actual].astype(str).str.strip()
                               if mun_actual and mun_actual in merged.columns
                               else pd.Series("", index=merged.index)),
        "ID_PARTICIPANTE":     _col(pid_svc).astype(str).str.strip(),
        "NOMBRE_PARTICIPANTE": _col(col_nombre_merged).astype(str).str.strip(),
        "TIPO_PARTICIPANTE":   _col(col_tipo_merged).astype(str).str.strip(),
        "TIPO_DOC":            _col(col_tdoc_merged).astype(str).str.strip(),
        "NUMERO_DOC":          _col(col_ndoc_merged).astype(str).str.strip(),
        "ID_SERVICIO":         _col(sid_svc).astype(str).str.strip(),
        "FECHA_ENTRADA_SERVICIO": _col(entrada_col),
        "PESTAÑA_ORIGEN":      pestana_origen,
        "OK_FILA":             ok_row.values,
        "N_PROBLEMAS":         n_problemas.values,
        "FALTANTES":           faltantes.values,
        "ERRORES_LOGICOS":     errores_serie.values,
        "ADVERTENCIAS":        advertencias_serie.values,
    }, index=merged.index)

    # Solo filas con problema o con advertencias
    tiene_problema = ~ok_row | advertencias_serie.ne("")
    casos = casos_df[tiene_problema].sort_values(
        ["OK_FILA", "N_PROBLEMAS"], ascending=[True, False]).copy()

    # Participantes huérfanos al final
    if len(orphans_part):
        dep_col_p  = dep_actual if dep_actual and dep_actual in orphans_part.columns else None
        mun_col_p  = mun_actual if mun_actual and mun_actual in orphans_part.columns else None
        nom_col_p  = col_nombre_p if col_nombre_p and col_nombre_p in orphans_part.columns else None
        tipo_col_p = tipo_part_col if tipo_part_col and tipo_part_col in orphans_part.columns else None
        tdoc_col_p = col_tipo_doc_p if col_tipo_doc_p and col_tipo_doc_p in orphans_part.columns else None
        ndoc_col_p = col_num_doc_p  if col_num_doc_p  and col_num_doc_p  in orphans_part.columns else None

        def _op(c):
            return (orphans_part[c].astype(str).str.strip() if c else "")

        orphan_rows = pd.DataFrame({
            "PROGRAMA":             "TOTAL_ARCHIVO",
            "DEPARTAMENTO":         _op(dep_col_p),
            "MUNICIPIO":            _op(mun_col_p),
            "ID_PARTICIPANTE":      orphans_part[pid_part].astype(str).str.strip(),
            "NOMBRE_PARTICIPANTE":  _op(nom_col_p),
            "TIPO_PARTICIPANTE":    _op(tipo_col_p),
            "TIPO_DOC":             _op(tdoc_col_p),
            "NUMERO_DOC":           _op(ndoc_col_p),
            "ID_SERVICIO":          "",
            "FECHA_ENTRADA_SERVICIO": pd.NaT,
            "PESTAÑA_ORIGEN":       "participantes",
            "OK_FILA":              False,
            "N_PROBLEMAS":          1,
            "FALTANTES":            "PARTICIPANTE HUÉRFANO: sin ningún servicio en el archivo",
            "ERRORES_LOGICOS":      "",
            "ADVERTENCIAS":         "",
        })
        casos = pd.concat([casos, orphan_rows], ignore_index=True)

    meta: Dict[str, Any] = {
        "periodo":                  periodo_str,
        "columna_filtro_periodo":   str(entrada_col),
        "registros_en_archivo":     str(n_archivo),
        "registros_en_periodo":     str(n_periodo),
        "registros_fuera_periodo":  str(n_fuera),
        "registros_sin_fecha":      str(int(svc["_ENTRADA_DT"].isna().sum())),
        "campos_evaluados":         str(len(all_fields)),
        "filas_con_problemas":      str(int((~ok_row).sum())),
        "filas_con_errores_logicos": str(int((~ok_logica).sum())),
        "filas_incompletas":        str(int((~ok_completitud).sum())),
        "n_participantes_huerfanos": str(len(orphans_part)),
        "n_dup_doc":                str(len(dup_doc_pids)),
        "n_dup_nombre":             str(len(dup_nombre_pids)),
        "n_dup_svc":                str(len(dup_svc_ids)),
        "pard_en_cohorte":          str(int(is_pard_merged.sum())),
        "archivo_entrada":          str(input_path),
    }

    return resumen_fila, resumen_celda, casos, meta


# ============================================================
# SECCIÓN 10 — ESCRITURA DEL REPORTE
# ============================================================

def write_output(
    path_out:      Path,
    resumen_fila:  pd.DataFrame,
    resumen_celda: pd.DataFrame,
    casos:         pd.DataFrame,
    progress_cb,
) -> None:
    """
    Escribe el Excel con tres pestañas — estructura idéntica a estadísticas_2025.
    casos_a_corregir:
      ROJO    = FALTANTES no vacío (campo sin diligenciar).
      NARANJA = FALTANTES vacío pero ERRORES_LOGICOS no vacío.
    """
    progress_cb("write", 0, 3, "Escribiendo pestaña: resumen_por_fila...")
    with pd.ExcelWriter(path_out, engine="openpyxl") as w:
        resumen_fila.to_excel(w,  index=False, sheet_name="resumen_por_fila")
        progress_cb("write", 1, 3, "Escribiendo pestaña: resumen_por_celda...")
        resumen_celda.to_excel(w, index=False, sheet_name="resumen_por_celda")
        progress_cb("write", 2, 3, "Escribiendo pestaña: casos_a_corregir...")
        casos.to_excel(w,         index=False, sheet_name="casos_a_corregir")
        progress_cb("write", 3, 3, "Aplicando formato y guardando...")

        for sname in ["resumen_por_fila", "resumen_por_celda", "casos_a_corregir"]:
            ws = w.sheets[sname]
            ws.freeze_panes    = "A2"
            ws.auto_filter.ref = ws.dimensions

        try:
            from openpyxl.styles import PatternFill
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.utils import get_column_letter

            ws_c     = w.sheets["casos_a_corregir"]
            header   = [cell.value for cell in ws_c[1]]
            last_col = get_column_letter(ws_c.max_column)
            last_row = ws_c.max_row
            rng      = f"A2:{last_col}{last_row}"

            fal_col = get_column_letter(header.index("FALTANTES") + 1)
            err_col = get_column_letter(header.index("ERRORES_LOGICOS") + 1)

            fill_rojo    = PatternFill(start_color="FFFFC7CE",
                                       end_color="FFFFC7CE", fill_type="solid")
            fill_naranja = PatternFill(start_color="FFFFE0B2",
                                       end_color="FFFFE0B2", fill_type="solid")

            # Rojo: hay campos vacíos (faltantes de completitud)
            ws_c.conditional_formatting.add(rng, FormulaRule(
                formula=[f'${fal_col}2<>""'], fill=fill_rojo))
            # Naranja: completitud OK pero hay error lógico
            ws_c.conditional_formatting.add(rng, FormulaRule(
                formula=[f'AND(${fal_col}2="",${err_col}2<>"")'],
                fill=fill_naranja))

        except Exception:
            pass


# ============================================================
# SECCIÓN 11 — PUNTO DE ENTRADA
# ============================================================

def main() -> None:
    root = tk.Tk()
    root.withdraw()

    # 1. Seleccionar período
    selector = PeriodSelectorDialog(root)
    if selector.result is None:
        root.destroy()
        return

    mes_ini, anio_ini, mes_fin, anio_fin = selector.result
    periodo_str = period_display(mes_ini, anio_ini, mes_fin, anio_fin)

    # 2. Seleccionar archivo
    p = filedialog.askopenfilename(
        title=f"Seleccione el archivo para el período: {periodo_str}",
        filetypes=[("Excel files", "*.xlsx;*.xls")],
    )
    if not p:
        root.destroy()
        return

    input_path = Path(p).resolve()
    period     = period_label(mes_ini, anio_ini, mes_fin, anio_fin)

    # 3. UI de progreso
    root.deiconify()
    ui = ProgressUI(root,
                    title=f"Calidad Participantes/Servicios — {periodo_str}")
    ui._log(f"Archivo  : {input_path}")
    ui._log(f"Período  : {periodo_str}")
    ui._log("")

    def worker() -> None:
        try:
            ui.progress("inicio", None, None,
                        f"Iniciando análisis para {periodo_str}...")
            resumen_fila, resumen_celda, casos, meta = run_checker(
                input_path, ui.progress,
                mes_ini, anio_ini, mes_fin, anio_fin,
            )
            out = default_output_path(input_path, period)
            write_output(out, resumen_fila, resumen_celda, casos, ui.progress)

            msg = (
                f"Período analizado    : {meta['periodo']}\n"
                f"Columna filtro       : {meta['columna_filtro_periodo']}\n"
                f"Registros en archivo : {meta['registros_en_archivo']}\n"
                f"Registros en período : {meta['registros_en_periodo']}\n"
                f"Fuera del período    : {meta['registros_fuera_periodo']}\n"
                f"Sin fecha válida     : {meta['registros_sin_fecha']}\n"
                f"Campos evaluados     : {meta['campos_evaluados']}\n"
                f"Filas con problemas  : {meta['filas_con_problemas']}\n"
                f"  · Incompletitud    : {meta['filas_incompletas']}\n"
                f"  · Errores lógicos  : {meta['filas_con_errores_logicos']}\n\n"
                f"Relacional:\n"
                f"  · Participantes huérfanos : {meta['n_participantes_huerfanos']}\n"
                f"  · Dup. doc participante   : {meta['n_dup_doc']}\n"
                f"  · Dup. nombre participante: {meta['n_dup_nombre']}\n"
                f"  · Dup. ID servicio        : {meta['n_dup_svc']}\n\n"
                f"Pestañas generadas:\n"
                f" - resumen_por_fila   : métricas a nivel de fila\n"
                f" - resumen_por_celda  : métricas a nivel de celda\n"
                f" - casos_a_corregir   : guía operativa\n"
                f"   · ROJO   = campos vacíos (completitud)\n"
                f"   · NARANJA = completa pero con error lógico\n"
                f"   · PESTAÑA_ORIGEN indica dónde ir a corregir"
            )
            ui.done(str(out), msg)

        except Exception as e:
            ui.error(str(e))

    threading.Thread(target=worker, daemon=True).start()
    ui.run()


if __name__ == "__main__":
    main()
